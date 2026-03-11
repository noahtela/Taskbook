import sys
from datetime import datetime, timedelta
from pathlib import Path

from PySide6.QtCore import QEvent, QObject, QThread, Qt, QTimer, Signal, Slot
from PySide6.QtGui import QAction, QColor, QKeySequence, QPalette, QShortcut
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QComboBox,
    QDialog,
    QFileDialog,
    QHeaderView,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QStyle,
    QSystemTrayIcon,
    QTableWidget,
    QTableWidgetItem,
    QToolBar,
    QMenu,
    QVBoxLayout,
    QWidget,
)
from openpyxl import Workbook, load_workbook

from app.repositories.ai_model_config_repository import AIModelConfigRepository
from app.repositories.prompt_template_repository import PromptTemplateRepository
from app.repositories.settings_repository import SettingsRepository
from app.repositories.task_repository import TaskRepository
from app.services.report_service import ReportService
from app.ui.ai_config_dialog import AIConfigDialog
from app.ui.prompt_config_dialog import PromptConfigDialog
from app.ui.report_preview_dialog import ReportPreviewDialog
from app.ui.settings_dialog import SettingsDialog
from app.ui.task_dialog import TaskDialog


STATUS_TEXT = {
    "todo": "待办",
    "doing": "进行中",
    "done": "已完成",
}


class ReportWorker(QObject):
    finished = Signal(str)
    failed = Signal(str)

    def __init__(self, report_service: ReportService, tasks, template, model_config):
        super().__init__()
        self.report_service = report_service
        self.tasks = tasks
        self.template = template
        self.model_config = model_config

    @Slot()
    def run(self):
        try:
            report_text = self.report_service.generate_daily_report(
                tasks=self.tasks,
                prompt_template=self.template,
                model_config=self.model_config,
            )
            self.finished.emit(report_text)
        except Exception as exc:
            self.failed.emit(str(exc))


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self._tray_enabled = False
        self.tray_icon = None
        self.setWindowTitle("Taskbook - 任务记事本")
        self.resize(860, 540)

        self.repo = TaskRepository()
        self.model_repo = AIModelConfigRepository()
        self.prompt_repo = PromptTemplateRepository()
        self.settings_repo = SettingsRepository()
        self.report_service = ReportService()

        self._has_user_sorted = False
        self._tray_enabled = QSystemTrayIcon.isSystemTrayAvailable()
        self._notified_due_task_ids = set()
        self._report_thread = None
        self._report_worker = None
        self._report_generating_dialog = None
        self._win10_toast_notifier = None

        self.search_timer = QTimer(self)
        self.search_timer.setSingleShot(True)
        self.search_timer.setInterval(300)
        self.search_timer.timeout.connect(self.refresh_table)

        self.reminder_timer = QTimer(self)
        self.reminder_timer.timeout.connect(self.check_upcoming_due_tasks)

        self.periodic_timer = QTimer(self)
        self.periodic_timer.timeout.connect(self.check_periodic_tasks)

        self._create_actions()
        self._build_menu_bar()
        # 工具栏功能已全部折叠到菜单栏，避免透明/系统主题下重叠，移除工具栏
        self.statusBar().showMessage("就绪")

        # 导入示例路径
        self.sample_import_path = str(Path.cwd() / "task_import_sample.xlsx")

        root = QWidget()
        root.setObjectName("centralwidget")
        self.setCentralWidget(root)

        layout = QVBoxLayout(root)

        filter_row = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("搜索标题/描述")
        self.search_input.setClearButtonEnabled(True)

        self.status_filter = QComboBox()
        self.status_filter.addItem("全部状态", "")
        self.status_filter.addItem("待办", "todo")
        self.status_filter.addItem("进行中", "doing")
        self.status_filter.addItem("已完成", "done")

        self.due_filter = QComboBox()
        self.due_filter.addItem("截止时间:不限", "all")
        self.due_filter.addItem("今天", "today")
        self.due_filter.addItem("本周", "week")
        self.due_filter.addItem("本月", "month")

        self.btn_search = QPushButton("搜索")

        filter_row.addWidget(self.search_input, 1)
        filter_row.addWidget(self.status_filter)
        filter_row.addWidget(self.due_filter)
        filter_row.addWidget(self.btn_search)
        layout.addLayout(filter_row)

        self.table = QTableWidget(0, 6)
        self.table.setHorizontalHeaderLabels(["ID", "标题", "状态", "优先级", "截止日期", "更新时间"])
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setColumnHidden(0, True)
        self.table.setSortingEnabled(True)
        self._setup_table_header()
        layout.addWidget(self.table)

        self.btn_search.clicked.connect(self.search_now)
        self.search_input.returnPressed.connect(self.search_now)
        self.search_input.textChanged.connect(self.schedule_auto_search)
        self.status_filter.currentIndexChanged.connect(self.refresh_table)
        self.due_filter.currentIndexChanged.connect(self.refresh_table)
        self.table.itemDoubleClicked.connect(self.edit_selected_task)
        self.table.horizontalHeader().sectionClicked.connect(self.mark_user_sorted)
        self.table.itemSelectionChanged.connect(self._update_action_states)

        self.shortcut_focus_search = QShortcut(QKeySequence("Ctrl+F"), self)
        self.shortcut_focus_search.activated.connect(self.focus_search)

        self.tray_icon = None
        if self._tray_enabled:
            self._create_tray_icon()

        self._init_windows_toast_notifier()
        self.apply_theme_from_settings()
        self._show_author_info_once_per_day()

        # 创建导入示例文件（如果不存在）
        self._ensure_import_sample()

        self.refresh_table()
        self._update_action_states()
        self._apply_reminder_settings()
        self.check_upcoming_due_tasks()

    def _setup_table_header(self):
        header = self.table.horizontalHeader()
        header.setStretchLastSection(True)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)

    def _create_actions(self):
        self.action_add = QAction("新建任务", self)
        self.action_add.setShortcut(QKeySequence("Ctrl+N"))
        self.action_add.setStatusTip("新建任务 (Ctrl+N)")
        self.action_add.setToolTip("新建任务 (Ctrl+N)")
        self.action_add.triggered.connect(self.add_task)

        self.action_done = QAction("批量完成", self)
        self.action_done.setShortcut(QKeySequence("Ctrl+D"))
        self.action_done.setStatusTip("将选中的任务标记为完成 (Ctrl+D)")
        self.action_done.setToolTip("批量完成 (Ctrl+D)")
        self.action_done.triggered.connect(self.mark_done)

        self.action_delete = QAction("批量删除", self)
        self.action_delete.setShortcut(QKeySequence("Delete"))
        self.action_delete.setStatusTip("删除选中的任务 (Delete)")
        self.action_delete.setToolTip("批量删除 (Delete)")
        self.action_delete.triggered.connect(self.delete_task)

        self.action_report = QAction("生成日报", self)
        self.action_report.setShortcut(QKeySequence("Ctrl+R"))
        self.action_report.setStatusTip("根据选中任务生成 AI 日报 (Ctrl+R)")
        self.action_report.setToolTip("生成日报 (Ctrl+R)")
        self.action_report.triggered.connect(self.generate_daily_report)

        self.action_export = QAction("导出为Excel", self)
        self.action_export.setShortcut(QKeySequence("Ctrl+Shift+E"))
        self.action_export.setStatusTip("导出任务到 Excel (.xlsx) (Ctrl+Shift+E)")
        self.action_export.setToolTip("导出任务到 Excel (.xlsx) (Ctrl+Shift+E)")
        self.action_export.triggered.connect(self.export_tasks_to_excel)

        self.action_import = QAction("导入任务", self)
        self.action_import.setShortcut(QKeySequence("Ctrl+Shift+I"))
        self.action_import.setStatusTip("从 Excel 导入任务 (Ctrl+Shift+I)")
        self.action_import.setToolTip("从 Excel 导入任务 (Ctrl+Shift+I)")
        self.action_import.triggered.connect(self.import_tasks_from_excel)

        self.action_save_import_sample = QAction("保存导入示例", self)
        self.action_save_import_sample.setShortcut(QKeySequence("Ctrl+Alt+I"))
        self.action_save_import_sample.setStatusTip("保存导入示例模板到本地 (Ctrl+Alt+I)")
        self.action_save_import_sample.setToolTip("保存导入示例模板到本地 (Ctrl+Alt+I)")
        self.action_save_import_sample.triggered.connect(self.save_import_sample)

        self.action_ai_config = QAction("AI模型配置", self)
        self.action_ai_config.setShortcut(QKeySequence("Ctrl+Shift+M"))
        self.action_ai_config.setStatusTip("打开 AI 模型配置 (Ctrl+Shift+M)")
        self.action_ai_config.setToolTip("AI模型配置 (Ctrl+Shift+M)")
        self.action_ai_config.triggered.connect(self.open_ai_config_dialog)

        self.action_prompt_config = QAction("Prompt配置", self)
        self.action_prompt_config.setShortcut(QKeySequence("Ctrl+Shift+P"))
        self.action_prompt_config.setStatusTip("打开 Prompt 配置 (Ctrl+Shift+P)")
        self.action_prompt_config.setToolTip("Prompt配置 (Ctrl+Shift+P)")
        self.action_prompt_config.triggered.connect(self.open_prompt_config_dialog)

        self.action_settings = QAction("系统设置", self)
        self.action_settings.setShortcut(QKeySequence("Ctrl+"))
        self.action_settings.setStatusTip("打开系统设置 (Ctrl+,)")
        self.action_settings.setToolTip("系统设置 (Ctrl+,)")
        self.action_settings.triggered.connect(self.open_settings_dialog)

        self.addAction(self.action_add)
        self.addAction(self.action_done)
        self.addAction(self.action_delete)
        self.addAction(self.action_report)
        self.addAction(self.action_export)
        self.addAction(self.action_import)
        self.addAction(self.action_ai_config)
        self.addAction(self.action_prompt_config)
        self.addAction(self.action_settings)

    def _build_menu_bar(self):
        task_menu = self.menuBar().addMenu("任务")
        task_menu.addAction(self.action_add)
        task_menu.addAction(self.action_done)
        task_menu.addAction(self.action_delete)
        task_menu.addSeparator()
        task_menu.addAction(self.action_report)
        task_menu.addAction(self.action_export)
        task_menu.addAction(self.action_import)
        task_menu.addAction(self.action_save_import_sample)

        ai_menu = self.menuBar().addMenu("AI")
        ai_menu.addAction(self.action_ai_config)
        ai_menu.addAction(self.action_prompt_config)

        settings_menu = self.menuBar().addMenu("设置")
        settings_menu.addAction(self.action_settings)


    def _create_tray_icon(self):
        self.tray_icon = QSystemTrayIcon(self)
        icon = self.windowIcon()
        if icon.isNull():
            style = QApplication.style()
            icon = style.standardIcon(QStyle.SP_DesktopIcon)
        self.tray_icon.setIcon(icon)
        self.tray_icon.setToolTip("Taskbook")

        tray_menu = QMenu(self)
        action_show = QAction("显示主窗口", self)
        action_show.triggered.connect(self.restore_from_tray)
        tray_menu.addAction(action_show)

        tray_menu.addSeparator()

        action_exit = QAction("退出", self)
        action_exit.triggered.connect(QApplication.instance().quit)
        tray_menu.addAction(action_exit)

        self.tray_icon.setContextMenu(tray_menu)
        self.tray_icon.activated.connect(self._on_tray_activated)
        self.tray_icon.show()

    def _on_tray_activated(self, reason):
        if reason == QSystemTrayIcon.DoubleClick:
            self.restore_from_tray()

    def restore_from_tray(self):
        # 清除最小化标志，确保可见并前置
        self.setWindowState((self.windowState() & ~Qt.WindowMinimized) | Qt.WindowActive)
        self.showNormal()
        self.raise_()
        self.activateWindow()

    def _update_action_states(self):
        has_selection = len(self.selected_task_ids()) > 0
        self.action_done.setEnabled(has_selection)
        self.action_delete.setEnabled(has_selection)
        self.action_report.setEnabled(has_selection and self._report_thread is None)

    def mark_user_sorted(self, _section: int):
        self._has_user_sorted = True

    def focus_search(self):
        self.search_input.setFocus()
        self.search_input.selectAll()

    def schedule_auto_search(self, _text: str):
        self.search_timer.start()

    def search_now(self):
        if self.search_timer.isActive():
            self.search_timer.stop()
        self.refresh_table()

    def _current_filters(self):
        keyword = self.search_input.text().strip()
        status = self.status_filter.currentData()
        due = self.due_filter.currentData()
        return keyword, status, due

    def refresh_table(self):
        keyword, status, due = self._current_filters()
        tasks = self.repo.list_tasks(keyword=keyword, status=status or None, due_filter=due)
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(tasks))

        now = datetime.now()
        for row, task in enumerate(tasks):
            self.table.setItem(row, 0, QTableWidgetItem(str(task.id)))
            self.table.setItem(row, 1, QTableWidgetItem(task.title))
            self.table.setItem(row, 2, QTableWidgetItem(STATUS_TEXT.get(task.status, task.status)))
            self.table.setItem(row, 3, QTableWidgetItem(str(task.priority)))
            self.table.setItem(row, 4, QTableWidgetItem(task.due_date or ""))
            self.table.setItem(row, 5, QTableWidgetItem(task.updated_at))

            if task.status == "done":
                for col in range(1, 6):
                    item = self.table.item(row, col)
                    if item is not None:
                        item.setForeground(Qt.gray)
            elif task.due_date:
                try:
                    due_dt = datetime.strptime(task.due_date, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    due_dt = None
                if due_dt is not None and due_dt < now:
                    due_item = self.table.item(row, 4)
                    if due_item is not None:
                        due_item.setForeground(Qt.red)

        self.table.setSortingEnabled(True)

        if self._has_user_sorted:
            header = self.table.horizontalHeader()
            self.table.sortItems(header.sortIndicatorSection(), header.sortIndicatorOrder())

        self._update_action_states()

    def selected_task_id(self):
        row = self.table.currentRow()
        if row < 0:
            return None
        item = self.table.item(row, 0)
        if item is None:
            return None
        return int(item.text())

    def selected_task_ids(self):
        model = self.table.selectionModel()
        if model is None:
            return []

        rows = sorted({index.row() for index in model.selectedRows()})
        task_ids = []
        for row in rows:
            item = self.table.item(row, 0)
            if item is not None:
                task_ids.append(int(item.text()))
        return task_ids

    def _current_filtered_tasks(self):
        keyword, status, due = self._current_filters()
        return self.repo.list_tasks(keyword=keyword, status=status or None, due_filter=due)

    def add_task(self):
        dialog = TaskDialog(self)
        if dialog.exec():
            payload = dialog.get_data()
            if not payload["title"]:
                QMessageBox.warning(self, "提示", "标题不能为空")
                return
            self.repo.create_task(
                title=payload["title"],
                description=payload["description"],
                status=payload["status"],
                priority=payload["priority"],
                due_date=payload["due_date"],
            )
            self.refresh_table()

    def edit_selected_task(self, _item=None):
        task_id = self.selected_task_id()
        if task_id is None:
            return

        task = self.repo.get_task(task_id)
        if task is None:
            QMessageBox.warning(self, "提示", "任务不存在，可能已被删除")
            self.refresh_table()
            return

        dialog = TaskDialog(self, task=task)
        if dialog.exec():
            payload = dialog.get_data()
            if not payload["title"]:
                QMessageBox.warning(self, "提示", "标题不能为空")
                return
            self.repo.update_task(
                task_id=task_id,
                title=payload["title"],
                description=payload["description"],
                status=payload["status"],
                priority=payload["priority"],
                due_date=payload["due_date"],
            )
            self.refresh_table()

    def mark_done(self):
        task_ids = self.selected_task_ids()
        if not task_ids:
            QMessageBox.information(self, "提示", "请先选择任务")
            return

        self.repo.mark_done_bulk(task_ids)
        self.refresh_table()

    def delete_task(self):
        task_ids = self.selected_task_ids()
        if not task_ids:
            QMessageBox.information(self, "提示", "请先选择任务")
            return

        count = len(task_ids)
        ok = QMessageBox.question(self, "确认", f"确定删除选中的 {count} 个任务吗？")
        if ok != QMessageBox.Yes:
            return

        self.repo.delete_tasks_bulk(task_ids)
        self.refresh_table()

    def open_ai_config_dialog(self):
        dialog = AIConfigDialog(self)
        dialog.exec()

    def open_prompt_config_dialog(self):
        dialog = PromptConfigDialog(self)
        dialog.exec()

    def open_settings_dialog(self):
        dialog = SettingsDialog(self)
        dialog.set_reminder_test_handler(self.trigger_due_reminder_test)
        if dialog.exec():
            dialog.save()
            self.apply_theme_from_settings()
            self._notified_due_task_ids.clear()
            self.check_upcoming_due_tasks()
            self.statusBar().showMessage("系统设置已保存", 3000)

    def apply_theme_from_settings(self):
        theme = self.settings_repo.get("theme", "light")
        alpha_str = self.settings_repo.get("transparent_alpha", "85")

        try:
            alpha_percent = min(100, max(50, int(alpha_str)))
        except ValueError:
            alpha_percent = 85

        app = QApplication.instance()
        if app is None:
            return

        self.setAttribute(Qt.WA_TranslucentBackground, False)

        if theme == "dark":
            palette = QPalette()
            palette.setColor(QPalette.Window, QColor("#1f1f1f"))
            palette.setColor(QPalette.WindowText, QColor("#f2f2f2"))
            palette.setColor(QPalette.Base, QColor("#2a2a2a"))
            palette.setColor(QPalette.AlternateBase, QColor("#333333"))
            palette.setColor(QPalette.Text, QColor("#f2f2f2"))
            palette.setColor(QPalette.Button, QColor("#2f2f2f"))
            palette.setColor(QPalette.ButtonText, QColor("#f2f2f2"))
            palette.setColor(QPalette.Highlight, QColor("#5a8cff"))
            palette.setColor(QPalette.HighlightedText, Qt.white)
            app.setPalette(palette)
            app.setStyleSheet(
                "QMainWindow { background-color: #1f1f1f; color: #f2f2f2; }"
                "QWidget#centralwidget { background-color: #1f1f1f; color: #f2f2f2; }"
                "QMenuBar { background-color: #262626; color: #f2f2f2; border-bottom: 1px solid #3a3a3a; }"
                "QMenuBar::item:selected { background: #343d4d; }"
                "QToolBar { background-color: #262626; border: none; border-bottom: 1px solid #3a3a3a; spacing: 6px; }"
                "QToolButton { background: transparent; color: #f2f2f2; padding: 4px 6px; border-radius: 4px; }"
                "QToolButton:hover { background: #343d4d; }"
                "QStatusBar { background-color: #262626; color: #f2f2f2; border-top: 1px solid #3a3a3a; }"
                "QLineEdit, QComboBox, QPushButton { background: #2f2f2f; color: #f2f2f2; border: 1px solid #4a4a4a; border-radius: 4px; min-height: 24px; }"
                "QPushButton:hover { background: #3a3a3a; }"
                "QTableWidget { background: #242424; color: #f2f2f2; gridline-color: #3d3d3d; border: 1px solid #3d3d3d; }"
                "QHeaderView::section { background: #303030; color: #f2f2f2; border: 0; border-right: 1px solid #3d3d3d; border-bottom: 1px solid #3d3d3d; padding: 4px; }"
                "QComboBox QAbstractItemView { background: #2f2f2f; color: #f2f2f2; selection-background-color: #5a8cff; selection-color: #ffffff; }"
                "QMenu { background: #2c2c2c; color: #f2f2f2; }"
                "QMenu::item:selected { background: #3a4b66; color: #ffffff; }"
            )
            self.setWindowOpacity(1.0)
        elif theme == "light":
            palette = QPalette()
            palette.setColor(QPalette.Window, QColor("#f6f7f9"))
            palette.setColor(QPalette.WindowText, QColor("#111111"))
            palette.setColor(QPalette.Base, QColor("#ffffff"))
            palette.setColor(QPalette.AlternateBase, QColor("#f8f8f8"))
            palette.setColor(QPalette.Text, QColor("#111111"))
            palette.setColor(QPalette.Button, QColor("#ffffff"))
            palette.setColor(QPalette.ButtonText, QColor("#111111"))
            palette.setColor(QPalette.Highlight, QColor("#5a8cff"))
            palette.setColor(QPalette.HighlightedText, Qt.white)
            app.setPalette(palette)
            app.setStyleSheet(
                "QMainWindow { background-color: #f6f7f9; color: #111111; }"
                "QWidget#centralwidget { background-color: #f6f7f9; color: #111111; }"
                "QMenuBar { background-color: #ffffff; color: #111111; border-bottom: 1px solid #d9d9d9; }"
                "QMenuBar::item:selected { background: #eaf2ff; }"
                "QToolBar { background-color: #ffffff; border: none; border-bottom: 1px solid #d9d9d9; spacing: 6px; }"
                "QToolButton { background: transparent; color: #111111; padding: 4px 6px; border-radius: 4px; }"
                "QToolButton:hover { background: #eaf2ff; }"
                "QStatusBar { background-color: #ffffff; color: #111111; border-top: 1px solid #d9d9d9; }"
                "QLineEdit, QComboBox, QPushButton { background: #ffffff; color: #111111; border: 1px solid #cfcfcf; border-radius: 4px; min-height: 24px; }"
                "QPushButton:hover { background: #f3f7ff; }"
                "QTableWidget { background: #ffffff; color: #111111; gridline-color: #e2e2e2; border: 1px solid #d7d7d7; }"
                "QHeaderView::section { background: #f3f3f3; color: #111111; border: 0; border-right: 1px solid #e1e1e1; border-bottom: 1px solid #e1e1e1; padding: 4px; }"
                "QComboBox QAbstractItemView { background: #ffffff; color: #111111; selection-background-color: #dbeafe; selection-color: #111111; }"
                "QMenu { background: #ffffff; color: #111111; }"
                "QMenu::item:selected { background: #eaf2ff; color: #111111; }"
            )
            self.setWindowOpacity(1.0)
        elif theme == "transparent":
            app.setPalette(app.style().standardPalette())
            app.setStyleSheet(
                "QMainWindow { background: transparent; }"
                "QWidget#centralwidget { background: transparent; }"
                "QTableWidget { background-color: rgba(255,255,255,220); }"
                "QLineEdit, QComboBox, QPushButton { background-color: rgba(255,255,255,220); color: #111111; }"
                "QComboBox QAbstractItemView { background: #ffffff; color: #111111; }"
                "QMenu { background: rgba(255,255,255,235); color: #111111; }"
                "QMenuBar { background-color: rgba(255,255,255,235); color: #111111; border-bottom: 1px solid rgba(0,0,0,40); }"
                "QMenuBar::item:selected { background: #eaf2ff; }"
                "QToolBar { background-color: rgba(255,255,255,235); border: none; border-bottom: 1px solid rgba(0,0,0,40); spacing: 6px; }"
                "QToolButton { background: transparent; color: #111111; padding: 4px 6px; border-radius: 4px; }"
                "QToolButton:hover { background: #eaf2ff; }"
                "QStatusBar { background-color: rgba(255,255,255,235); color: #111111; border-top: 1px solid rgba(0,0,0,40); }"
            )
            self.setAttribute(Qt.WA_TranslucentBackground, True)
            self.setWindowOpacity(alpha_percent / 100.0)
        else:
            app.setPalette(app.style().standardPalette())
            app.setStyleSheet("")
            self.setWindowOpacity(1.0)

    def _apply_reminder_settings(self):
        remind_enabled = self.settings_repo.get("remind_enabled", "1") == "1"
        threshold_str = self.settings_repo.get("remind_threshold_minutes", "60")
        periodic_enabled = self.settings_repo.get("periodic_enabled", "0") == "1"
        periodic_minutes_str = self.settings_repo.get("periodic_minutes", "60")

        # 到期提醒
        if remind_enabled:
            try:
                threshold_minutes = max(1, int(threshold_str))
            except ValueError:
                threshold_minutes = 60
            self.reminder_timer.setInterval(threshold_minutes * 60 * 1000)
            self.reminder_timer.start()
        else:
            self.reminder_timer.stop()

        # 定期提醒
        if periodic_enabled:
            try:
                periodic_minutes = max(15, int(periodic_minutes_str))
            except ValueError:
                periodic_minutes = 60
            self.periodic_timer.setInterval(periodic_minutes * 60 * 1000)
            self.periodic_timer.start()
        else:
            self.periodic_timer.stop()

    def check_upcoming_due_tasks(self):
        remind_enabled = self.settings_repo.get("remind_enabled", "1") == "1"
        if not remind_enabled:
            return
        threshold_str = self.settings_repo.get("remind_threshold_minutes", "60")
        try:
            threshold_minutes = max(1, int(threshold_str))
        except ValueError:
            threshold_minutes = 60

        now = datetime.now()
        threshold_time = now + timedelta(minutes=threshold_minutes)

        tasks = self.repo.list_tasks(status=None)
        upcoming = []
        for task in tasks:
            if task.status == "done" or not task.due_date:
                continue
            try:
                due_dt = datetime.strptime(task.due_date, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue

            if now <= due_dt <= threshold_time and task.id not in self._notified_due_task_ids:
                upcoming.append(task)

        if not upcoming:
            return

        top = upcoming[0]
        total = len(upcoming)
        title = "任务即将到期提醒"
        message = f"{top.title} 将在阈值内到期"
        if total > 1:
            message += f"，另有 {total - 1} 个任务。"

        self.show_windows_notification(title, message)

        for task in upcoming:
            if task.id is not None:
                self._notified_due_task_ids.add(task.id)

    def show_windows_notification(self, title: str, message: str):
        try:
            dialog = QDialog(self)
            dialog.setWindowTitle(title)
            dialog.setModal(True)
            dialog.setFixedSize(360, 180)

            layout = QVBoxLayout(dialog)
            layout.setContentsMargins(16, 16, 16, 16)
            layout.setSpacing(12)

            label = QLabel(message)
            label.setWordWrap(True)
            layout.addWidget(label)

            btn_row = QHBoxLayout()
            btn_row.addStretch()
            btn_close = QPushButton("关闭")
            btn_close.clicked.connect(dialog.accept)
            btn_row.addWidget(btn_close)
            layout.addLayout(btn_row)

            dialog.show()
            dialog.raise_()
            dialog.activateWindow()
        except Exception:
            QMessageBox.information(self, title, message)

    def _init_windows_toast_notifier(self):
        # 保留原逻辑以兼容未来需求，但当前提醒已改为程序内弹窗
        self._win10_toast_notifier = None

    def _show_author_info_once_per_day(self):
        last_shown = self.settings_repo.get("author_info_last_shown", "")
        today = datetime.now().strftime("%Y-%m-%d")
        if last_shown == today:
            return
        msg = (
            "作者：卿云团队-yangsir\n"
            "联系方式：noah704523@gmail.com (欢迎反馈)\n"
            "感谢使用，祝工作顺利！"
        )
        QMessageBox.information(self, "关于作者", msg)
        self.settings_repo.set("author_info_last_shown", today)

    def _show_windows_toast_best_effort(self, title: str, message: str) -> bool:
        return False

    def trigger_due_reminder_test(self):
        self.show_windows_notification("任务即将到期提醒(测试)", "这是一条固定测试提醒：系统通知功能正常。")

    def generate_daily_report(self):
        task_ids = self.selected_task_ids()
        if not task_ids:
            QMessageBox.information(self, "提示", "请先选择任务")
            return

        if self._report_thread is not None:
            QMessageBox.information(self, "提示", "已有日报生成任务在进行中，请稍候")
            return

        model_config = self.model_repo.get_active_config()
        if model_config is None:
            QMessageBox.warning(self, "提示", "请先配置并启用 AI 模型")
            return

        template = self.prompt_repo.get_active_template(scene="daily_report")
        if template is None:
            QMessageBox.warning(self, "提示", "请先配置并启用日报 Prompt 模板")
            return

        tasks = self.repo.list_tasks_by_ids(task_ids)
        if not tasks:
            QMessageBox.warning(self, "提示", "未找到选中的任务")
            return

        self._set_report_actions_enabled(False)
        self.statusBar().showMessage("正在生成日报...")

        if self._report_generating_dialog is None:
            self._report_generating_dialog = QDialog(self)
            self._report_generating_dialog.setWindowTitle("正在生成日报")
            self._report_generating_dialog.setModal(True)
            self._report_generating_dialog.setFixedSize(320, 140)
            layout = QVBoxLayout(self._report_generating_dialog)
            layout.setContentsMargins(16, 16, 16, 16)
            layout.setSpacing(12)
            label = QLabel("正在生成日报，请稍候...")
            label.setWordWrap(True)
            layout.addWidget(label)
        self._report_generating_dialog.show()
        self._report_generating_dialog.raise_()
        self._report_generating_dialog.activateWindow()

        self._report_thread = QThread(self)
        self._report_worker = ReportWorker(self.report_service, tasks, template, model_config)
        self._report_worker.moveToThread(self._report_thread)

        self._report_thread.started.connect(self._report_worker.run)
        self._report_worker.finished.connect(self._on_report_generated)
        self._report_worker.failed.connect(self._on_report_failed)
        self._report_worker.finished.connect(self._cleanup_report_thread)
        self._report_worker.failed.connect(self._cleanup_report_thread)
        self._report_thread.start()

    def _set_report_actions_enabled(self, enabled: bool):
        self.action_report.setEnabled(enabled and len(self.selected_task_ids()) > 0)

    @Slot(str)
    def _on_report_generated(self, report_text: str):
        self.statusBar().showMessage("日报生成完成", 3000)
        self._close_report_generating_dialog()
        preview = ReportPreviewDialog(report_text, self)
        preview.exec()

    @Slot(str)
    def _on_report_failed(self, error_text: str):
        self.statusBar().showMessage("日报生成失败", 3000)
        self._close_report_generating_dialog()
        QMessageBox.warning(self, "生成失败", error_text)

    @Slot()
    def _cleanup_report_thread(self):
        self._set_report_actions_enabled(True)
        self._close_report_generating_dialog()

        if self._report_thread is not None:
            self._report_thread.quit()
            self._report_thread.wait()
            self._report_thread.deleteLater()
        if self._report_worker is not None:
            self._report_worker.deleteLater()

        self._report_thread = None
        self._report_worker = None

    def _close_report_generating_dialog(self):
        if self._report_generating_dialog is not None:
            try:
                self._report_generating_dialog.accept()
            except Exception:
                self._report_generating_dialog.close()
            self._report_generating_dialog = None

    def check_periodic_tasks(self):
        periodic_enabled = self.settings_repo.get("periodic_enabled", "0") == "1"
        if not periodic_enabled:
            return
        tasks = self.repo.list_tasks(status=None)
        pending = [t for t in tasks if t.status in ("todo", "doing")]
        if not pending:
            return
        title = "定期提醒"
        message = f"共有 {len(pending)} 个待办/进行中任务，请及时查看。"
        self.show_windows_notification(title, message)

    def export_tasks_to_excel(self):
        selection = QMessageBox.question(
            self,
            "导出范围",
            "选择导出范围：\n是 -> 导出当前筛选结果\n否 -> 仅导出选中任务",
            QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
            QMessageBox.Yes,
        )
        if selection == QMessageBox.Cancel:
            return

        if selection == QMessageBox.Yes:
            tasks = self._current_filtered_tasks()
        else:
            ids = self.selected_task_ids()
            if not ids:
                QMessageBox.information(self, "提示", "请先选择要导出的任务")
                return
            tasks = self.repo.list_tasks_by_ids(ids)

        if not tasks:
            QMessageBox.information(self, "提示", "没有可导出的任务")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "保存为 Excel",
            "tasks.xlsx",
            "Excel 文件 (*.xlsx)",
        )
        if not file_path:
            return

        status_text = {
            "todo": "待办",
            "doing": "进行中",
            "done": "已完成",
        }
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "任务列表"
            headers = ["标题", "状态", "优先级", "截止日期", "更新时间", "描述"]
            ws.append(headers)
            for task in tasks:
                ws.append(
                    [
                        task.title,
                        status_text.get(task.status, task.status),
                        task.priority,
                        task.due_date or "",
                        task.updated_at,
                        task.description or "",
                    ]
                )

            # 示例模板另存一份，方便用户导入参考
            try:
                wb.save(self.sample_import_path)
            except Exception:
                pass

            wb.save(file_path)
            QMessageBox.information(self, "完成", f"已导出到 {file_path}")
        except Exception as exc:
            QMessageBox.warning(self, "导出失败", str(exc))

    def import_tasks_from_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择要导入的 Excel",
            "",
            "Excel 文件 (*.xlsx)",
        )
        if not file_path:
            return

        required_headers = ["标题", "状态", "优先级", "截止日期", "更新时间", "描述"]
        status_map = {
            "待办": "todo",
            "进行中": "doing",
            "已完成": "done",
        }

        try:
            wb = load_workbook(file_path)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if headers != required_headers:
                QMessageBox.warning(
                    self,
                    "导入失败",
                    "表头不匹配，请使用示例模板：标题、状态、优先级、截止日期、更新时间、描述",
                )
                return

            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                title, status_cn, priority, due_date, updated_at, description = row
                # 校验必填
                if not (title and status_cn and priority and due_date and updated_at and description):
                    QMessageBox.warning(self, "导入失败", "存在必填字段为空，请检查数据后重试")
                    return
                status = status_map.get(status_cn)
                if status is None:
                    QMessageBox.warning(self, "导入失败", f"状态值无效：{status_cn}，请使用 待办/进行中/已完成")
                    return
                try:
                    priority_int = int(priority)
                    if priority_int not in (1, 2, 3):
                        raise ValueError()
                except Exception:
                    QMessageBox.warning(self, "导入失败", "优先级必须是 1/2/3")
                    return

                rows.append(
                    {
                        "title": str(title),
                        "description": str(description),
                        "status": status,
                        "priority": priority_int,
                        "due_date": self._normalize_datetime_str(due_date),
                        "created_at": self._normalize_datetime_str(updated_at),
                        "updated_at": self._normalize_datetime_str(updated_at),
                        "completed_at": self._normalize_datetime_str(updated_at) if status == "done" else None,
                    }
                )

            if not rows:
                QMessageBox.information(self, "提示", "没有可导入的数据")
                return

            inserted = self.repo.bulk_import(rows)
            QMessageBox.information(self, "完成", f"成功导入 {inserted} 条任务")
            self.refresh_table()
        except Exception as exc:
            QMessageBox.warning(self, "导入失败", str(exc))

    def save_import_sample(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "保存导入示例",
            "task_import_sample.xlsx",
            "Excel 文件 (*.xlsx)",
        )
        if not file_path:
            return
        try:
            self._write_import_sample(Path(file_path))
            QMessageBox.information(self, "完成", f"已保存示例到 {file_path}")
        except Exception as exc:
            QMessageBox.warning(self, "保存失败", str(exc))

    def _normalize_datetime_str(self, value):
        if value is None:
            return ""
        try:
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d %H:%M:%S")
            return str(value)
        except Exception:
            return str(value)

    def _ensure_import_sample(self):
        try:
            path = Path(self.sample_import_path)
            if path.exists():
                return
            self._write_import_sample(path)
        except Exception:
            pass

    def _write_import_sample(self, path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "任务列表"
        headers = ["标题", "状态", "优先级", "截止日期", "更新时间", "描述"]
        ws.append(headers)
        ws.append(["示例任务", "待办", 2, "2026-12-31 18:00:00", "2026-01-01 10:00:00", "示例描述"])
        wb.save(path)

    def closeEvent(self, event):
        if self._tray_enabled and self.tray_icon is not None:
            box = QMessageBox(self)
            box.setWindowTitle("退出确认")
            box.setText("请选择操作：")
            exit_btn = box.addButton("退出程序", QMessageBox.YesRole)
            tray_btn = box.addButton("最小化到托盘", QMessageBox.NoRole)
            box.addButton("取消", QMessageBox.RejectRole)
            box.exec()
            clicked = box.clickedButton()
            if clicked == exit_btn:
                event.accept()
                return
            if clicked == tray_btn:
                event.ignore()
                self.hide()
                self.tray_icon.showMessage("Taskbook", "程序已最小化到托盘", QSystemTrayIcon.Information, 3000)
                return
            event.ignore()
            return

        reply = QMessageBox.question(
            self,
            "退出确认",
            "确定要退出程序吗？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()

    def changeEvent(self, event):
        super().changeEvent(event)
        if not self._tray_enabled or self.tray_icon is None:
            return
        if event.type() == QEvent.WindowStateChange and self.isMinimized():
            self.hide()
            self.tray_icon.showMessage("Taskbook", "程序已最小化到托盘", QSystemTrayIcon.Information, 3000)
